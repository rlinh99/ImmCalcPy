from openpyxl import load_workbook
from .classes import *
from datetime import datetime


def read_workbook(wb):
    sheet = wb.active
    data = sheet.iter_rows()
    return data


def insert_data(rows):
    data = []

    for row in rows:
        data.append(Record(row[0].value, row[1].value, row[3].value))

    return [x for x in data if x.date is not None]


def calculate_date(data, given_date):
    index, current = get_current(data)
    date_pairs = []

    for record in data[index:]:
        # only consider the case in Canada
        if record.location != "Canada":
            continue

        # handle new in and out cycle
        if record.type == "In":
            current = record
            continue

        # handle current cycle
        if (current is not None) and (record.type == "Out"):
            date_pairs.append({"In": current.date, "Out": record.date})
            current = None

    date_pairs.append({"In": current.date, "Out": given_date})

    return get_result(date_pairs)


def get_result(data):
    result = 0
    last_out = None
    for item in data:
        placeholder = 1
        if item["In"] == last_out:
            placeholder = 0

        duration = (item["Out"] - item["In"]).days + placeholder
        result += duration
        last_out = item["Out"]

    return result


def get_current(data):
    for index, record in enumerate(data):
        if (record.type == "In") and (record.location == "Canada"):
            return index, record

    return -1, None


def get_given_date():
    while True:
        input_date = input("Please enter given date in format 'mm/dd/yyyy':  ")
        if validate_date(input_date):
            return datetime.strptime(input_date, '%m/%d/%Y')


def validate_date(d):
    try:
        if len(d) == 10:
            datetime.strptime(d, '%m/%d/%Y')
            return True
        else:
            print("Input is Invalid, please input again")
            return False

    except ValueError:
        print("Input is Invliad, please input again")
        return False


def get_file():
    while True:
        try:
            filename = input("Please enter the .xlsx data file")
            # return load_workbook("test.xlsx")
            return load_workbook(filename)

        except FileNotFoundError:
            print("File is not found, please check your file path and enter again")
